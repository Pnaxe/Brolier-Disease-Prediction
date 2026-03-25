"""
Poultry Disease Classifier - FastAPI Backend
Run: uvicorn app:app --reload --port 8000
"""
import os
import sys
from io import BytesIO
from copy import copy
import threading
import subprocess
import re
import sqlite3
from datetime import datetime, timezone

import numpy as np
from fastapi import FastAPI, File, Form, UploadFile, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from PIL import Image
from pydantic import BaseModel
from openpyxl import load_workbook

if sys.version_info >= (3, 14):
    raise RuntimeError(
        "This backend requires Python 3.11, 3.12, or 3.13. TensorFlow is not available for "
        f"Python {sys.version_info.major}.{sys.version_info.minor} in this setup."
    )

try:
    import tensorflow as tf
except ImportError as exc:
    raise RuntimeError(
        "TensorFlow is not installed. Use Python 3.11, 3.12, or 3.13 and run "
        "`python -m pip install -r requirements.txt` inside `server/`."
    ) from exc

# Paths
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATASET_PATH = r"C:\Users\USER\Pictures\Final Project\poultry_diseases"
MODEL_PATH = os.path.join(BASE_DIR, "poultry_model.keras")
# Fallback: check parent folder if model was trained from old Gradio script
MODEL_PATH_FALLBACK = os.path.join(os.path.dirname(BASE_DIR), "poultry_model.keras")
REPORT_TEMPLATE_PATH = os.path.join(BASE_DIR, "Broiler_Disease_Report.xlsx")
TRAIN_SCRIPT_PATH = os.path.join(BASE_DIR, "train.py")
TRAINING_DATASET_ROOT = os.path.join(BASE_DIR, "training_datasets", "current")
MODELS_DIR = os.path.join(BASE_DIR, "models")
ACTIVE_MODEL_FILE = os.path.join(MODELS_DIR, "active_model.txt")
DB_PATH = os.path.join(BASE_DIR, "app_data.sqlite3")
CONFIDENCE_THRESHOLD = 0.65

app = FastAPI(title="Poultry Disease Classifier API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173", "http://127.0.0.1:5173", "http://localhost:3000"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Model state
model = None
model_ready = False
active_model_name = None
class_names = ["cocci", "healthy", "ncd", "salmo"]  # alphabetical
img_size = 224
training_state = {
    "running": False,
    "action": None,
    "status": "idle",
    "last_started_at": None,
    "last_finished_at": None,
    "last_exit_code": None,
    "last_message": "No training run yet.",
    "log_tail": [],
    "dataset_path": DATASET_PATH,
    "active_model_name": None,
    "current_epoch": 0,
    "total_epochs": 0,
    "progress_percent": 0.0,
    "progress_phase": "idle",
    "train_accuracy": None,
    "validation_accuracy": None,
    "train_loss": None,
    "validation_loss": None,
    "cancel_requested": False,
}
training_lock = threading.Lock()
current_training_process = None
progress_pattern = re.compile(
    r"^TRAIN_PROGRESS\|phase=(?P<phase>[^|]+)\|epoch=(?P<epoch>\d+)\|total=(?P<total>\d+)\|progress=(?P<progress>[\d.]+)\|"
    r"train_acc=(?P<train_acc>[\d.]+)\|val_acc=(?P<val_acc>[\d.]+)\|train_loss=(?P<train_loss>[\d.]+)\|val_loss=(?P<val_loss>[\d.]+)$"
)

seed_batches = [
    {
        "name": "Batch A - 500 birds",
        "size": 500,
        "age": "2 weeks",
        "location": "House 1",
        "farm": "Newcastle North Farm",
        "stage": "Starter",
        "status": "Healthy",
        "archived": 0,
        "linked_cases_count": 1,
        "last_check": "Today, 08:45",
    },
    {
        "name": "Layer Unit C",
        "size": 1980,
        "age": "12 weeks",
        "location": "Unit C",
        "farm": "Green Valley Poultry",
        "stage": "Layers",
        "status": "Watchlist",
        "archived": 0,
        "linked_cases_count": 1,
        "last_check": "Today, 07:10",
    },
    {
        "name": "Block D",
        "size": 2840,
        "age": "6 weeks",
        "location": "Block D",
        "farm": "Riverbend Broilers",
        "stage": "Finisher",
        "status": "Priority",
        "archived": 0,
        "linked_cases_count": 2,
        "last_check": "Today, 06:30",
    },
    {
        "name": "Starter Pen 2",
        "size": 1120,
        "age": "2 weeks",
        "location": "Pen 2",
        "farm": "Sunrise Layers Unit",
        "stage": "Starter",
        "status": "Healthy",
        "archived": 1,
        "linked_cases_count": 0,
        "last_check": "Yesterday",
    },
]

seed_cases = [
    {"title": "Possible NCD exposure", "priority": "Medium", "status": "Pending review", "batch_name": "Batch A - 500 birds"},
    {"title": "Respiratory distress review", "priority": "Medium", "status": "Lab follow-up", "batch_name": "Layer Unit C"},
    {"title": "High mortality alert", "priority": "High", "status": "Vet escalations", "batch_name": "Block D"},
    {"title": "Feed drop investigation", "priority": "Medium", "status": "Treatment plans", "batch_name": "Block D"},
    {"title": "Coccidiosis lab follow-up", "priority": "Low", "status": "Resolved this week", "batch_name": "Starter Pen 2"},
]


def count_dataset_images(dataset_root: str):
    summary = {}
    for label in class_names:
        class_dir = os.path.join(dataset_root, label)
        if not os.path.isdir(class_dir):
            summary[label] = 0
            continue
        summary[label] = len(
            [name for name in os.listdir(class_dir) if name.lower().endswith((".png", ".jpg", ".jpeg"))]
        )
    return summary


def ensure_model_dir():
    os.makedirs(MODELS_DIR, exist_ok=True)


def read_active_model_name():
    ensure_model_dir()
    if not os.path.exists(ACTIVE_MODEL_FILE):
        return None
    with open(ACTIVE_MODEL_FILE, "r", encoding="utf-8") as handle:
        value = handle.read().strip()
    return value or None


def write_active_model_name(model_name: str):
    ensure_model_dir()
    with open(ACTIVE_MODEL_FILE, "w", encoding="utf-8") as handle:
        handle.write(model_name)


def list_available_models():
    ensure_model_dir()
    models = []
    for name in sorted(os.listdir(MODELS_DIR)):
        if not name.lower().endswith(".keras"):
            continue
        path = os.path.join(MODELS_DIR, name)
        if os.path.isfile(path):
            models.append({"name": name, "path": path, "active": name == active_model_name})
    return models


def sanitize_model_name(model_name: str):
    cleaned = "".join(ch if ch.isalnum() or ch in {"-", "_"} else "_" for ch in model_name.strip())
    cleaned = cleaned.strip("_")
    return cleaned or "model"


def get_db_connection():
    connection = sqlite3.connect(DB_PATH)
    connection.row_factory = sqlite3.Row
    return connection


def init_db():
    with get_db_connection() as connection:
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS batches (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE,
                size INTEGER NOT NULL,
                age TEXT NOT NULL,
                location TEXT NOT NULL,
                farm TEXT NOT NULL,
                stage TEXT NOT NULL,
                status TEXT NOT NULL,
                archived INTEGER NOT NULL DEFAULT 0,
                linked_cases_count INTEGER NOT NULL DEFAULT 0,
                last_check TEXT,
                created_at TEXT NOT NULL
            )
            """
        )
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS cases (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                title TEXT NOT NULL,
                priority TEXT NOT NULL,
                status TEXT NOT NULL,
                batch_name TEXT,
                created_at TEXT NOT NULL
            )
            """
        )
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS predictions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                file_name TEXT NOT NULL,
                batch_name TEXT,
                disease TEXT NOT NULL,
                confidence REAL NOT NULL,
                low_confidence INTEGER NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL
            )
            """
        )
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS reports (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                report_name TEXT NOT NULL,
                report_type TEXT NOT NULL,
                created_at TEXT NOT NULL
            )
            """
        )
        connection.commit()


def seed_db():
    now = datetime.now(timezone.utc).isoformat()
    with get_db_connection() as connection:
        existing_batch_count = connection.execute("SELECT COUNT(*) AS total FROM batches").fetchone()["total"]
        if existing_batch_count == 0:
            connection.executemany(
                """
                INSERT INTO batches (name, size, age, location, farm, stage, status, archived, linked_cases_count, last_check, created_at)
                VALUES (:name, :size, :age, :location, :farm, :stage, :status, :archived, :linked_cases_count, :last_check, :created_at)
                """,
                [{**row, "created_at": now} for row in seed_batches],
            )

        existing_case_count = connection.execute("SELECT COUNT(*) AS total FROM cases").fetchone()["total"]
        if existing_case_count == 0:
            connection.executemany(
                """
                INSERT INTO cases (title, priority, status, batch_name, created_at)
                VALUES (:title, :priority, :status, :batch_name, :created_at)
                """,
                [{**row, "created_at": now} for row in seed_cases],
            )
        connection.commit()


def log_prediction(file_name: str, batch_name: str | None, disease: str, confidence: float, low_confidence: bool):
    with get_db_connection() as connection:
        connection.execute(
            """
            INSERT INTO predictions (file_name, batch_name, disease, confidence, low_confidence, created_at)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (
                file_name,
                batch_name.strip() if batch_name else None,
                disease,
                confidence,
                1 if low_confidence else 0,
                datetime.now(timezone.utc).isoformat(),
            ),
        )
        connection.commit()


def iso_to_display(iso_value: str):
    try:
        parsed = datetime.fromisoformat(iso_value)
    except ValueError:
        return iso_value
    return parsed.astimezone().strftime("%d %b, %H:%M")


def build_dashboard_summary():
    with get_db_connection() as connection:
        total_screened = connection.execute("SELECT COUNT(*) AS total FROM predictions").fetchone()["total"]
        screened_today = connection.execute(
            "SELECT COUNT(*) AS total FROM predictions WHERE DATE(created_at) = DATE('now')"
        ).fetchone()["total"]
        alerts_total = connection.execute(
            "SELECT COUNT(*) AS total FROM predictions WHERE low_confidence = 0 AND disease != 'healthy'"
        ).fetchone()["total"]
        alerts_today = connection.execute(
            "SELECT COUNT(*) AS total FROM predictions WHERE DATE(created_at) = DATE('now') AND low_confidence = 0 AND disease != 'healthy'"
        ).fetchone()["total"]
        active_batches = connection.execute(
            "SELECT COUNT(*) AS total FROM batches WHERE archived = 0"
        ).fetchone()["total"]
        watchlist_batches = connection.execute(
            "SELECT COUNT(*) AS total FROM batches WHERE archived = 0 AND status = 'Watchlist'"
        ).fetchone()["total"]
        priority_batches = connection.execute(
            "SELECT COUNT(*) AS total FROM batches WHERE archived = 0 AND status = 'Priority'"
        ).fetchone()["total"]
        healthy_batches = connection.execute(
            "SELECT COUNT(*) AS total FROM batches WHERE archived = 0 AND status = 'Healthy'"
        ).fetchone()["total"]
        active_cases = connection.execute(
            "SELECT COUNT(*) AS total FROM cases WHERE status != 'Resolved this week'"
        ).fetchone()["total"]
        active_farms = connection.execute(
            "SELECT COUNT(DISTINCT farm) AS total FROM batches WHERE archived = 0"
        ).fetchone()["total"]
        average_confidence = connection.execute(
            "SELECT AVG(confidence) AS average_confidence FROM predictions WHERE low_confidence = 0"
        ).fetchone()["average_confidence"]

        workflow_statuses = ["Pending review", "Lab follow-up", "Vet escalations", "Treatment plans", "Resolved this week"]
        workflow_rows = []
        for status_name in workflow_statuses:
            total = connection.execute(
                "SELECT COUNT(*) AS total FROM cases WHERE status = ?",
                (status_name,),
            ).fetchone()["total"]
            workflow_rows.append({"label": status_name, "value": str(total)})

        recent_prediction_rows = connection.execute(
            """
            SELECT batch_name, disease, created_at
            FROM predictions
            WHERE low_confidence = 0
            ORDER BY datetime(created_at) DESC
            LIMIT 3
            """
        ).fetchall()

        if recent_prediction_rows:
            insight_rows = [
                {
                    "title": f"{row['batch_name'] or 'Unassigned batch'} flagged for {str(row['disease']).upper()}",
                    "meta": iso_to_display(row["created_at"]),
                }
                for row in recent_prediction_rows
            ]
        else:
            recent_cases = connection.execute(
                """
                SELECT title, batch_name, created_at
                FROM cases
                ORDER BY datetime(created_at) DESC
                LIMIT 3
                """
            ).fetchall()
            insight_rows = [
                {
                    "title": f"{row['title']} recorded for {row['batch_name'] or 'general review'}",
                    "meta": iso_to_display(row["created_at"]),
                }
                for row in recent_cases
            ]

        def daily_counts(days: int, disease_filter: str | None = None):
            values = []
            for day_offset in range(days - 1, -1, -1):
                if disease_filter is None:
                    query = """
                        SELECT COUNT(*) AS total FROM predictions
                        WHERE DATE(created_at) = DATE('now', ?)
                    """
                    params = (f"-{day_offset} day",)
                else:
                    query = """
                        SELECT COUNT(*) AS total FROM predictions
                        WHERE DATE(created_at) = DATE('now', ?)
                        AND low_confidence = 0
                        AND disease != 'healthy'
                    """
                    params = (f"-{day_offset} day",)
                values.append(connection.execute(query, params).fetchone()["total"])
            return values

        screening_volume = daily_counts(7)
        prediction_trend = daily_counts(8)
        alert_trend = daily_counts(8, disease_filter="alerts")

    healthy_ratio = round((healthy_batches / active_batches) * 100, 1) if active_batches else 0.0
    watchlist_ratio = round((watchlist_batches / active_batches) * 100, 1) if active_batches else 0.0
    priority_ratio = round((priority_batches / active_batches) * 100, 1) if active_batches else 0.0

    return {
        "model_ready": model_ready,
        "hero_tags": [
            f"{active_farms} farms active",
            f"{screened_today} droppings screenings today",
            f"{priority_batches} critical batches flagged",
        ],
        "hero_side": {
            "suspected_cases": f"{active_cases} active cases",
            "detection_model": "Ready" if model_ready else "Offline",
        },
        "overview_cards": [
            {
                "label": "Droppings Images Screened",
                "value": f"{total_screened:,}",
                "change": f"+{screened_today} today",
                "trend": "up",
                "icon": "scan",
            },
            {
                "label": "Suspected Disease Alerts",
                "value": f"{alerts_total:,}",
                "change": f"+{alerts_today} new flagged batches",
                "trend": "up",
                "icon": "alert",
            },
            {
                "label": "Batches Under Observation",
                "value": f"{watchlist_batches + priority_batches:,}",
                "change": f"{priority_batches} need immediate review",
                "trend": "up",
                "icon": "case",
            },
            {
                "label": "Healthy Batches",
                "value": f"{healthy_batches:,}",
                "change": f"{healthy_ratio}% of active batches",
                "trend": "up",
                "icon": "ai",
            },
        ],
        "quick_stats": [
            {"value": f"{active_farms}", "label": "Active farms"},
            {"value": f"{screened_today}", "label": "Samples reviewed today"},
            {"value": f"{round(average_confidence or 0)}%", "label": "Model confidence average"},
        ],
        "risk_levels": [
            {"label": "Healthy batches", "value": f"{healthy_batches}", "width": f"{max(healthy_ratio, 8)}%"},
            {"label": "Watchlist batches", "value": f"{watchlist_batches}", "width": f"{max(watchlist_ratio, 8)}%"},
            {"label": "Critical batches", "value": f"{priority_batches}", "width": f"{max(priority_ratio, 8)}%"},
        ],
        "workflow_rows": workflow_rows,
        "insight_rows": insight_rows,
        "screening_volume": screening_volume,
        "prediction_trend": prediction_trend,
        "alert_trend": alert_trend,
    }


class ReportRow(BaseModel):
    caseId: str
    flock: str
    date: str
    disease: str
    confidence: int
    status: str
    recommendation: str


class ReportPayload(BaseModel):
    rows: list[ReportRow]


class TrainingRequest(BaseModel):
    action: str
    dataset_path: str | None = None
    model_name: str | None = None


class ActiveModelRequest(BaseModel):
    model_name: str


def build_model():
    m = tf.keras.Sequential([
        tf.keras.layers.Input(shape=(img_size, img_size, 3)),
        tf.keras.layers.Conv2D(32, 3, activation="relu"),
        tf.keras.layers.BatchNormalization(),
        tf.keras.layers.Dropout(0.2),
        tf.keras.layers.MaxPooling2D(),
        tf.keras.layers.Conv2D(64, 3, activation="relu"),
        tf.keras.layers.BatchNormalization(),
        tf.keras.layers.Dropout(0.25),
        tf.keras.layers.MaxPooling2D(),
        tf.keras.layers.Conv2D(128, 3, activation="relu"),
        tf.keras.layers.BatchNormalization(),
        tf.keras.layers.Dropout(0.3),
        tf.keras.layers.MaxPooling2D(),
        tf.keras.layers.Conv2D(256, 3, activation="relu"),
        tf.keras.layers.BatchNormalization(),
        tf.keras.layers.Dropout(0.3),
        tf.keras.layers.MaxPooling2D(),
        tf.keras.layers.GlobalAveragePooling2D(),
        tf.keras.layers.Dense(256, activation="relu"),
        tf.keras.layers.Dropout(0.5),
        tf.keras.layers.Dense(len(class_names), activation="softmax"),
    ])
    m.compile(optimizer="adam", loss="sparse_categorical_crossentropy", metrics=["accuracy"])
    return m


def load_model():
    global model, model_ready, active_model_name
    model = build_model()
    active_model_name = read_active_model_name()
    training_state["active_model_name"] = active_model_name
    if not active_model_name:
        print("No active model configured.")
        model_ready = False
        return

    path = os.path.join(MODELS_DIR, active_model_name)
    if not os.path.exists(path):
        print(f"Configured model not found: {path}")
        model_ready = False
        return

    try:
        model.load_weights(path)
        model_ready = True
        training_state["active_model_name"] = active_model_name
        print(f"Model loaded from {path}")
    except Exception as e:
        print(f"Could not load model: {e}")
        model_ready = False


def append_training_log(line: str):
    if not line:
        return

    match = progress_pattern.match(line)
    if match:
        training_state["current_epoch"] = int(match.group("epoch"))
        training_state["total_epochs"] = int(match.group("total"))
        training_state["progress_percent"] = float(match.group("progress"))
        training_state["progress_phase"] = match.group("phase")
        training_state["train_accuracy"] = float(match.group("train_acc"))
        training_state["validation_accuracy"] = float(match.group("val_acc"))
        training_state["train_loss"] = float(match.group("train_loss"))
        training_state["validation_loss"] = float(match.group("val_loss"))
        phase = training_state["progress_phase"]
        if phase == "training":
            training_state["last_message"] = (
                f"Epoch {training_state['current_epoch']} of {training_state['total_epochs']} running."
            )
        elif phase == "prepare":
            training_state["last_message"] = "Preparing dataset and training pipeline."
        elif phase == "complete":
            training_state["last_message"] = "Training finished and model is being saved."
        return

    logs = training_state["log_tail"]
    logs.append(line)
    training_state["log_tail"] = logs[-20:]


def suspend_process(pid: int):
    subprocess.run(
        ["powershell", "-Command", f"Suspend-Process -Id {pid}"],
        check=True,
        capture_output=True,
        text=True,
    )


def resume_process(pid: int):
    subprocess.run(
        ["powershell", "-Command", f"Resume-Process -Id {pid}"],
        check=True,
        capture_output=True,
        text=True,
    )


def cancel_process(pid: int):
    subprocess.run(
        ["taskkill", "/PID", str(pid), "/T", "/F"],
        check=True,
        capture_output=True,
        text=True,
    )


def run_training_job(action: str, dataset_path: str, model_name: str):
    global current_training_process
    with training_lock:
        training_state["running"] = True
        training_state["action"] = action
        training_state["status"] = "running"
        training_state["last_started_at"] = datetime.now(timezone.utc).isoformat()
        training_state["last_finished_at"] = None
        training_state["last_exit_code"] = None
        training_state["last_message"] = f"{action.title()} started."
        training_state["dataset_path"] = dataset_path
        training_state["log_tail"] = [f"{action.title()} started at {training_state['last_started_at']}"]
        training_state["current_epoch"] = 0
        training_state["total_epochs"] = 0
        training_state["progress_percent"] = 0.0
        training_state["progress_phase"] = "starting"
        training_state["train_accuracy"] = None
        training_state["validation_accuracy"] = None
        training_state["train_loss"] = None
        training_state["validation_loss"] = None
        training_state["cancel_requested"] = False

    try:
        ensure_model_dir()
        model_output_name = f"{sanitize_model_name(model_name)}.keras"
        model_output_path = os.path.join(MODELS_DIR, model_output_name)
        env = os.environ.copy()
        env["DATASET_PATH_OVERRIDE"] = dataset_path
        env["MODEL_OUTPUT_PATH"] = model_output_path
        env["PYTHONUNBUFFERED"] = "1"
        process = subprocess.Popen(
            [sys.executable, "-u", TRAIN_SCRIPT_PATH],
            cwd=BASE_DIR,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            env=env,
        )
        current_training_process = process

        if process.stdout:
            for line in process.stdout:
                append_training_log(line.rstrip())

        exit_code = process.wait()

        with training_lock:
            current_training_process = None
            training_state["running"] = False
            training_state["last_finished_at"] = datetime.now(timezone.utc).isoformat()
            training_state["last_exit_code"] = exit_code

            if training_state["cancel_requested"]:
                training_state["status"] = "canceled"
                training_state["progress_phase"] = "canceled"
                training_state["last_message"] = "Training canceled."
                append_training_log(training_state["last_message"])
            elif exit_code == 0:
                write_active_model_name(model_output_name)
                load_model()
                training_state["status"] = "completed"
                training_state["progress_percent"] = 100.0
                training_state["progress_phase"] = "completed"
                training_state["last_message"] = f"{action.title()} completed successfully and model reloaded."
                training_state["active_model_name"] = model_output_name
                append_training_log(training_state["last_message"])
            else:
                training_state["status"] = "failed"
                training_state["progress_phase"] = "failed"
                training_state["last_message"] = f"{action.title()} failed. Check the log output."
                append_training_log(training_state["last_message"])
    except Exception as exc:
        with training_lock:
            current_training_process = None
            training_state["running"] = False
            training_state["status"] = "failed"
            training_state["progress_phase"] = "failed"
            training_state["last_finished_at"] = datetime.now(timezone.utc).isoformat()
            training_state["last_message"] = f"{action.title()} failed to start: {exc}"
            append_training_log(training_state["last_message"])


@app.on_event("startup")
def startup():
    init_db()
    seed_db()
    load_model()


@app.get("/")
def root():
    return {"message": "Poultry Disease Classifier API", "model_ready": model_ready}


@app.get("/status")
def status():
    return {"model_ready": model_ready, "classes": class_names}


@app.get("/dashboard/summary")
def dashboard_summary():
    return build_dashboard_summary()


@app.get("/training/status")
def training_status():
    active_model_path = os.path.join(MODELS_DIR, active_model_name) if active_model_name else None
    return {
        **training_state,
        "model_ready": model_ready,
        "available_models": list_available_models(),
        "model_path": active_model_path,
        "confidence_threshold": CONFIDENCE_THRESHOLD,
        "dataset_summary": count_dataset_images(TRAINING_DATASET_ROOT),
        "uploaded_dataset_path": TRAINING_DATASET_ROOT,
    }


@app.get("/training/models")
def training_models():
    return {
        "active_model_name": active_model_name,
        "models": list_available_models(),
        "model_ready": model_ready,
    }


@app.post("/training/active-model")
def training_active_model(payload: ActiveModelRequest):
    model_name = payload.model_name.strip()
    if not model_name:
        raise HTTPException(status_code=400, detail="Model name is required.")

    path = os.path.join(MODELS_DIR, model_name)
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail="Selected model was not found.")

    write_active_model_name(model_name)
    load_model()
    training_state["active_model_name"] = model_name
    return {"message": f"{model_name} is now the active model.", "model_ready": model_ready}


@app.post("/training/upload-dataset/{label}")
async def training_upload_dataset(label: str, files: list[UploadFile] = File(...), clear_existing: bool = True):
    normalized_label = label.lower().strip()
    if normalized_label not in class_names:
        raise HTTPException(status_code=400, detail=f"Label must be one of: {', '.join(class_names)}")

    class_dir = os.path.join(TRAINING_DATASET_ROOT, normalized_label)
    os.makedirs(class_dir, exist_ok=True)

    if clear_existing:
        for existing_name in os.listdir(class_dir):
            existing_path = os.path.join(class_dir, existing_name)
            if os.path.isfile(existing_path):
                os.remove(existing_path)

    saved_count = 0
    for upload in files:
        if not upload.filename:
            continue
        if not upload.content_type or not upload.content_type.startswith("image/"):
            continue

        safe_name = os.path.basename(upload.filename)
        destination = os.path.join(class_dir, safe_name)
        contents = await upload.read()
        with open(destination, "wb") as target:
            target.write(contents)
        saved_count += 1

    if saved_count == 0:
        raise HTTPException(status_code=400, detail="No valid image files were uploaded for this disease folder.")

    training_state["dataset_path"] = TRAINING_DATASET_ROOT
    return {
        "message": f"Uploaded {saved_count} image(s) to {normalized_label}.",
        "dataset_path": TRAINING_DATASET_ROOT,
        "dataset_summary": count_dataset_images(TRAINING_DATASET_ROOT),
    }


@app.delete("/training/upload-dataset/{label}")
def training_clear_uploaded_dataset(label: str):
    normalized_label = label.lower().strip()
    if normalized_label not in class_names:
        raise HTTPException(status_code=400, detail=f"Label must be one of: {', '.join(class_names)}")

    class_dir = os.path.join(TRAINING_DATASET_ROOT, normalized_label)
    os.makedirs(class_dir, exist_ok=True)

    removed_count = 0
    for existing_name in os.listdir(class_dir):
        existing_path = os.path.join(class_dir, existing_name)
        if os.path.isfile(existing_path):
            os.remove(existing_path)
            removed_count += 1

    training_state["dataset_path"] = TRAINING_DATASET_ROOT
    return {
        "message": f"Cleared {removed_count} image(s) from {normalized_label}.",
        "dataset_path": TRAINING_DATASET_ROOT,
        "dataset_summary": count_dataset_images(TRAINING_DATASET_ROOT),
    }


@app.post("/training/start")
def training_start(payload: TrainingRequest):
    action = payload.action.lower().strip()
    if action not in {"train", "recalibrate"}:
        raise HTTPException(status_code=400, detail="Action must be 'train' or 'recalibrate'.")

    dataset_path = (payload.dataset_path or "").strip()
    if not dataset_path:
        raise HTTPException(status_code=400, detail="Dataset path is required.")
    if not os.path.exists(dataset_path):
        raise HTTPException(status_code=400, detail="Dataset path does not exist.")
    if not os.path.isdir(dataset_path):
        raise HTTPException(status_code=400, detail="Dataset path must be a folder.")

    model_name = (payload.model_name or "").strip()
    if not model_name:
        raise HTTPException(status_code=400, detail="Model name is required.")

    model_output_name = f"{sanitize_model_name(model_name)}.keras"
    model_output_path = os.path.join(MODELS_DIR, model_output_name)
    if os.path.exists(model_output_path):
        raise HTTPException(status_code=409, detail="A model with that name already exists. Choose a different name.")

    if training_state["running"]:
        raise HTTPException(status_code=409, detail="A training job is already running.")

    worker = threading.Thread(target=run_training_job, args=(action, dataset_path, model_name), daemon=True)
    worker.start()
    return {"accepted": True, "message": f"{action.title()} job started."}


@app.post("/training/pause")
def training_pause():
    if not training_state["running"] or current_training_process is None or current_training_process.poll() is not None:
        raise HTTPException(status_code=409, detail="No training job is currently running.")
    if training_state["status"] == "paused":
        raise HTTPException(status_code=409, detail="Training is already paused.")

    try:
        suspend_process(current_training_process.pid)
        training_state["status"] = "paused"
        training_state["progress_phase"] = "paused"
        training_state["last_message"] = "Training paused."
        append_training_log(training_state["last_message"])
        return {"message": "Training paused."}
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Could not pause training: {exc}")


@app.post("/training/resume")
def training_resume():
    if not training_state["running"] or current_training_process is None or current_training_process.poll() is not None:
        raise HTTPException(status_code=409, detail="No paused training job is available.")
    if training_state["status"] != "paused":
        raise HTTPException(status_code=409, detail="Training is not paused.")

    try:
        resume_process(current_training_process.pid)
        training_state["status"] = "running"
        training_state["progress_phase"] = "training"
        training_state["last_message"] = "Training resumed."
        append_training_log(training_state["last_message"])
        return {"message": "Training resumed."}
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Could not resume training: {exc}")


@app.post("/training/cancel")
def training_cancel():
    if not training_state["running"] or current_training_process is None or current_training_process.poll() is not None:
        raise HTTPException(status_code=409, detail="No training job is currently running.")

    try:
        training_state["cancel_requested"] = True
        training_state["status"] = "canceling"
        training_state["progress_phase"] = "canceling"
        training_state["last_message"] = "Cancel requested."
        append_training_log(training_state["last_message"])
        cancel_process(current_training_process.pid)
        return {"message": "Training cancel requested."}
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Could not cancel training: {exc}")


@app.post("/training/reload")
def training_reload():
    load_model()
    return {"message": "Model weights reloaded.", "model_ready": model_ready}


@app.post("/predict")
async def predict(file: UploadFile = File(...), batch_name: str | None = Form(None)):
    if not model_ready or model is None:
        raise HTTPException(status_code=503, detail="Model not loaded. Run train.py first.")

    if not file.content_type or not file.content_type.startswith("image/"):
        raise HTTPException(status_code=400, detail="File must be an image (PNG, JPG, JPEG)")

    try:
        contents = await file.read()
        img = Image.open(BytesIO(contents)).convert("RGB")
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid image: {str(e)}")

    arr = tf.keras.preprocessing.image.img_to_array(img)
    arr = tf.image.resize(arr, (img_size, img_size))
    arr = np.expand_dims(arr.numpy() if hasattr(arr, "numpy") else arr, 0) / 255.0

    preds = model.predict(arr, verbose=0)[0]
    idx = int(np.argmax(preds))
    conf = float(preds[idx])

    probabilities = {class_names[i]: round(float(preds[i]) * 100, 1) for i in range(len(class_names))}

    if conf < CONFIDENCE_THRESHOLD:
        response = {
            "disease": "Unable to classify",
            "confidence": round(conf * 100, 1),
            "low_confidence": True,
            "message": "This image does not appear to be poultry-related or the model is uncertain. Please upload a clear poultry image.",
            "probabilities": probabilities,
        }
        log_prediction(file.filename or "unnamed-image", batch_name, response["disease"], response["confidence"], True)
        return response

    response = {
        "disease": class_names[idx],
        "confidence": round(conf * 100, 1),
        "low_confidence": False,
        "message": None,
        "probabilities": probabilities,
    }
    log_prediction(file.filename or "unnamed-image", batch_name, response["disease"], response["confidence"], False)
    return response


@app.post("/report-excel")
def report_excel(payload: ReportPayload):
    if not os.path.exists(REPORT_TEMPLATE_PATH):
        raise HTTPException(status_code=500, detail="Report template not found on the server.")

    workbook = load_workbook(REPORT_TEMPLATE_PATH)
    worksheet = workbook["Disease Report"] if "Disease Report" in workbook.sheetnames else workbook.active

    template_row_index = 3
    start_row_index = 3
    template_cells = [worksheet.cell(template_row_index, column) for column in range(1, 8)]

    max_existing_row = max(worksheet.max_row, start_row_index)
    if max_existing_row >= start_row_index:
        worksheet.delete_rows(start_row_index, max_existing_row - start_row_index + 1)

    for row_offset, row in enumerate(payload.rows, start=start_row_index):
        worksheet.insert_rows(row_offset)
        values = [
            row.caseId,
            row.flock,
            row.date,
            row.disease,
            row.confidence,
            row.status,
            row.recommendation,
        ]

        for column_index, value in enumerate(values, start=1):
            cell = worksheet.cell(row_offset, column_index)
            template_cell = template_cells[column_index - 1]
            cell.value = value
            if template_cell.has_style:
                cell._style = copy(template_cell._style)
            if template_cell.number_format:
                cell.number_format = template_cell.number_format
            if template_cell.font:
                cell.font = copy(template_cell.font)
            if template_cell.fill:
                cell.fill = copy(template_cell.fill)
            if template_cell.border:
                cell.border = copy(template_cell.border)
            if template_cell.alignment:
                cell.alignment = copy(template_cell.alignment)
            if template_cell.protection:
                cell.protection = copy(template_cell.protection)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="Broiler_Disease_Report.xlsx"'},
    )


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
